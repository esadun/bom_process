import openpyxl
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinterdnd2 import TkinterDnD, DND_FILES


def process_file(file_path):
    try:
        BOMIdx=3 #0
        QtyIdx=7 #3
        SheetIdx=0
        print("Running, going to load workbook...")
        select_button["state"]="disabled"
        #root.update()
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        print (wb.sheetnames)
        print("...loaded.")
        #sheet = wb.active
        sheets = wb.sheetnames
        sheet = wb[sheets[SheetIdx]]
        print("...opened sheet ",SheetIdx)
        print("cols: ",sheet.max_column,", rows:",sheet.max_row)
        print("********************")
        for row in sheet.iter_rows(min_col=1, max_col=sheet.max_column, min_row=1, max_row=1):
            for cellIdx, cell in enumerate(row):
                if cellIdx==BOMIdx:
                    print("[BOM-->]"+str(cell.value or ""))
                elif cellIdx==QtyIdx:
                    print("[QTY-->]"+str(cell.value or ""))
                else:
                    print("          "+str(cell.value or ""))
        print("********************")
        print("Now going to process.")
        nestedBeyond2 = True
        while nestedBeyond2:
            nestedFound = 0
            nestedBeyond2 = False
            ancestorCellIndex = 0
            ancestorCellMultiplier = 1
            previousCellMultiplier = 1
            previousBOMlvl = 1
            maxRow=sheet.max_row
            for row in sheet.iter_rows(min_col=1, max_col=(max(BOMIdx,QtyIdx)+1), min_row=1, max_row=sheet.max_row):
                print("On row ",row[0].row, "(BOM=",row[BOMIdx].value,")")
                BOMlvl = row[BOMIdx]
                if BOMlvl.value is not None and isinstance(BOMlvl.value, (int, float)):
                    print("         Numeric BOM")
                    if BOMlvl.row==maxRow and BOMlvl.value > 2:
                        print("         Very bottom of sheet with BOM in need of reduction")
                        #print("AT BOTTOM, SEE BOM:",BOMlvl.value)
                    if BOMlvl.value > previousBOMlvl:  # at a top
                        print("         A2")
                        ancestorCellIndex = BOMlvl.row - 1
                        ancestorCellMultiplier = previousCellMultiplier
                    elif (BOMlvl.value < previousBOMlvl and previousBOMlvl > 2):  # at a bottom
                        print("         A3")
                        nestedFound += 1
                        nestedBeyond2 = True
                        print("         Going to loop from ",ancestorCellIndex + 1," to ",BOMlvl.row - 1)
                        for subRow in sheet.iter_rows(min_row=ancestorCellIndex + 1, max_row=BOMlvl.row - 1, min_col=1, max_col=(max(BOMIdx,QtyIdx)+1)):
                        #for subRow in sheet.iter_rows(min_row=ancestorCellIndex + 1, max_row=BOMlvl.row, min_col=1, max_col=(max(BOMIdx,QtyIdx)+1)):
                            print("         A.Inner looping, on sub-row ",subRow[0].row)
                            for subCell in subRow:
                                #print("                  sub-cell ",subCell.col_idx)
                                if subCell.col_idx == (BOMIdx+1):
                                    print("                  modifing BOM from ",subCell.value," to ",subCell.value - 1)
                                    subCell.value = subCell.value - 1  # BOM LEVEL                        
                                elif subCell.col_idx == (QtyIdx+1):
                                    print("                  modifing QTY from ",subCell.value," to ",subCell.value * ancestorCellMultiplier)
                                    subCell.value = subCell.value * ancestorCellMultiplier  # QtyPer
                        break
                    previousBOMlvl = BOMlvl.value
                    previousCellMultiplier = row[QtyIdx].value
                else:
                    print("         Non-numeric BOM")
                    if previousBOMlvl > 2:  # at a bottom
                        print("         B1")
                        nestedFound += 1
                        nestedBeyond2 = True
                        print("         Going to loop from ",ancestorCellIndex + 1," to ",BOMlvl.row - 1)
                        for subRow in sheet.iter_rows(min_row=ancestorCellIndex + 1, max_row=BOMlvl.row - 1, min_col=1, max_col=(max(BOMIdx,QtyIdx)+1)):
                        #for subRow in sheet.iter_rows(min_row=ancestorCellIndex + 1, max_row=BOMlvl.row, min_col=1, max_col=(max(BOMIdx,QtyIdx)+1)):
                            print("         B.Inner looping, on sub-row ",subRow[0].row)
                            for subCell in subRow:
                                print("                  sub-cell ",subCell.col_idx)
                                if subCell.col_idx == (BOMIdx+1):
                                    print("                  modifing BOM from ",subCell.value," to ",subCell.value - 1)
                                    subCell.value = subCell.value - 1  # BOM LEVEL                        
                                elif subCell.col_idx == (QtyIdx+1):
                                    print("                  modifing QTY from ",subCell.value," to ",subCell.value * ancestorCellMultiplier)
                                    subCell.value = subCell.value * ancestorCellMultiplier  # QtyPer
                        break
                    previousBOMlvl = 1

            #print("IVE FOUND: ", nestedFound)

        # Save the modified file
        modified_file_path = file_path.replace(".xlsx", "_modified.xlsx")
        wb.save(modified_file_path)
        messagebox.showinfo("Success", f"File processed and saved as: {modified_file_path}")
        select_button["state"]="normal"
        print("...done.")
        sys.exit()
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


def on_drag_and_drop(event):
    # Get the dropped file path
    file_path = event.data.strip()  # Removes extra whitespace or braces
    if file_path.startswith("{") and file_path.endswith("}"):
        file_path = file_path[1:-1]
    process_file(file_path)


def select_file():
    file_path = filedialog.askopenfilename(
        title="Select an Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    if file_path:
        process_file(file_path)

root = TkinterDnD.Tk()
root.title("Excel Processor with Drag-and-Drop")
root.geometry("400x200")
select_button = tk.Button(root, text="Select or Drop File", command=select_file, padx=20, pady=10)
select_button.pack()

def main():
    # Create a drag-and-drop-enabled Tkinter window
    #root = TkinterDnD.Tk()
    #root.title("Excel Processor with Drag-and-Drop")
    #root.geometry("400x200")

    # Enable drag-and-drop
    root.drop_target_register(DND_FILES)
    root.dnd_bind('<<Drop>>', on_drag_and_drop)

    

    root.mainloop()


if __name__ == "__main__":
    main()

'''
#go from top to bottom
    if current cell > previous cell, ancestor = current cell-1; get index
    else if current cell < previous cell and current cell -1 > 2, are at a bottom
        ancestor index + 1-->current cell -1
            val*=ancestor
            BOM--
'''